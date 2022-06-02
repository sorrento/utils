import os

from u_base import inicia, tardado


def getmtime(filename):
    """Return the last modification time of a file, reported by os.stat()."""
    return os.stat(filename).st_mtime


def getatime(filename):
    """Return the last access time of a file, reported by os.stat()."""
    return os.stat(filename).st_atime


def getctime(filename):
    """Return the metadata change time of a file, reported by os.stat()."""
    return os.stat(filename).st_ctime


def lista_files_recursiva(path, ext):
    """
devuelve la lista de archivos en la ruta, recursivamente, de la extensión especificada. la lista está ordenada por fecha
de modificación
    :param path:
    :param ext:
    :return:
    """
    import glob
    lista = glob.glob(path + '**/*.' + ext, recursive=True)
    lista = sorted(lista, key=getmtime, reverse=True)

    return lista


def fecha_mod(file):
    """
entrega la fecha de modificación de un archivo como un número yyyymmdd
    :param file:
    :return:
    """
    import datetime
    dt = datetime.datetime.fromtimestamp(getmtime(file))
    return int(dt.strftime('%Y%m%d'))


def get_filename(path):
    """
obtiene el nombre del fichero desde el path completo
    :param path:
    :return:
    """
    return os.path.basename(path)


def lee_txt(file_path):
    """
lee fichero de texto
    :param file_path:
    :return:
    """
    import os
    if os.path.isfile(file_path):
        # open text file in read mode
        text_file = open(file_path, "r", encoding='utf-8')

        # read whole file to a string
        data = text_file.read()

        # close file
        text_file.close()
        return data


def lee_excel(path_template):
    from openpyxl import load_workbook
    l = inicia('Leyendo:{}'.format(path_template))
    wb = load_workbook(path_template, keep_links=False, keep_vba=False)
    tardado(l)

    return wb


def remueve_sheet(wb, name):
    wb.remove(name)


def existe_sheet(wb, name):
    names = wb.sheetnames
    ex = name in names
    print('existe {}?{}'.format(name, ex))
    return ex


def escribe_sheet(wb, sh_name, df, sobreescribe=False):
    from openpyxl.utils.dataframe import dataframe_to_rows
    li = inicia('Escribiendo sheet')
    if existe_sheet(wb, sh_name):
        print('ya existe la sheet: {}'.format(sh_name))
        if sobreescribe:
            print('sobreescribiendo {}'.format(sh_name))
            sh = wb[sh_name]
            remueve_sheet(wb, sh)
        else:
            print('No sobreescribimos. Poner sobreescribir=True')
            return

    print('creamos la sheet:{}'.format(sh_name))
    sh = wb.create_sheet(sh_name)

    for r in dataframe_to_rows(df, index=False, header=True):
        sh.append(r)

    tardado(li)


def write_excel(wb, path, bucket):
    """

    Parameters
    ----------
    wb
    path
    bucket: bucket de S3. si es None se guarda en la ruta el local, (para pruebas)
    """
    import boto3
    from tempfile import NamedTemporaryFile

    li = inicia('Escribiendo {}'.format(path))

    if bucket is not None:
        s3 = boto3.resource('s3')
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            s3.meta.client.upload_file(tmp.name, bucket, path)
    else:
        print('** Guardando en local (no S3) porque se ha usado como parámetro bucket=None')
        wb.save(path)

    tardado(li)


def escribe_pestana(bucket, path_template, pestana, data, path_out):
    wb = lee_excel(path_template)
    escribe_sheet(wb, pestana, data, True)
    write_excel(wb, path_out, bucket)


def escribe_dos_pestanas(bucket, path_template, pestana_1, data_1, pestana_2, data_2, path_out):
    wb = lee_excel(path_template)
    escribe_sheet(wb, pestana_1, data_1, True)
    escribe_sheet(wb, pestana_2, data_2, True)
    write_excel(wb, path_out, bucket)
